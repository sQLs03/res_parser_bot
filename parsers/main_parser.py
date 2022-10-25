import logging
import os.path
import time
import zipfile
import tarfile

import pandas as pd
from pandas.io.excel import ExcelWriter

from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.types import input_file

from loader import bot
from config import PATH_TO_RESULT_REPORT
from parsers.joined_sell import parse_joined_sell
from parsers.EMA import parse_EMA_comment
from parsers.drops_detection import parse_DropsDetection_comment
from parsers.moon_hook_plus_spreaddetection import parse_hook_long_depth_comment, parse_moonstrike_comment
from parsers.moonshot import parse_moonshot
from parsers.pump_plus_strike_detection import parse_pumpdetection_plus_strikedetection
from parsers.puvp_detection import parse_PuvpDetection
from parsers.stread_detection import parse_StreadDetection
from parsers.creat_3_list import get_mean_plot
from parsers.waves import parse_waves_comment
from parsers.ichi_trend_bot import parse_ichi_trend_bot
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils import get_column_letter



async def get_headers(file_path: os.path, message: types.Message, state: FSMContext):
    """ Пробегается по файлу и для каждой строки записывает стратегию.

    :param file_path: Путь к файлу, который обрабатываем.
    :param message: Сообщение от пользователя
    :return: [strategy_1, strategy_2, ...]
    """
    result = []
    name_strategy=[]
    unknown_comments = set()
    with open(file_path, 'r') as file:
        first_str = file.readline()  # Это строка Binance Report - она не нужна
        if first_str.split()[:1] != ['Binance']:
            await message.answer(f"Файл не соответствует шаблону.", reply_markup=types.ReplyKeyboardRemove())
            await message.answer(f"Ошибка все еще не исправлена? - тогда вы можете обратиться к @NickByvaltsev",
                                 reply_markup=types.ReplyKeyboardRemove())
            await state.finish()
            return None
        headers = [header.strip() for header in file.readline().rstrip().split('\t')]
        try:
            comment_index = headers.index('Comment')
        except ValueError:
            await message.answer(f"Файл не соответствует шаблону, отсутствует строки Comment.", reply_markup=types.ReplyKeyboardRemove())
            await message.answer(f"Ошибка все еще не исправлена? - тогда вы можете обратиться к @NickByvaltsev",
                                 reply_markup=types.ReplyKeyboardRemove())
            await state.finish()
            return None
        for ind, line in enumerate(file.readlines()):
            line = line.rstrip().split('\t')

            try:
                comment = line[comment_index].lstrip()
            except IndexError:
                continue
            # Получаем стратегию и оставшийся комментарий
            if comment[:60].find(';') != -1:
                type_strategy, comment = comment.split(';', maxsplit=1)
            elif comment[:60].find(':') != -1:
                type_strategy, comment = comment.split(':', maxsplit=1)
            elif comment[:60].find('trend') != -1:
                type_strategy, comment = comment.split(maxsplit=1)[0], comment
            else:
                result.append('')
                len_before_adding = len(unknown_comments)
                unknown_comments.add(comment[:100])
                len_after_adding = len(unknown_comments)
                if len_before_adding != len_after_adding:
                    try:
                        await message.answer(f"Не могу распознать комментарий на {ind + 3} строке входного файла:\n\n"
                                             f"{comment[:30]}...", reply_markup=types.ReplyKeyboardRemove())
                    except:
                        await message.answer(f"Произошла ошибка при поиске названий стратегий. Возможно, "
                                             f"какая-то стратегия не была распознана.\n\n"
                                             f"Ошибка на {ind + 3} строке файла.",
                                             reply_markup=types.ReplyKeyboardRemove())
                continue
            result.append(type_strategy)
            if comment[:20].find("<") != -1:
                name_strategy.append(comment[comment.find("<"):comment.find(">") + 1])
            elif comment[:50].find(" (") != -1:
                name_strategy.append(comment[:50].split(" (")[0].split()[-1])
            elif comment[:50].find("using strategy") != -1:
                name_strategy.append("#" + comment.split("using strategy ")[0].split()[0])
            else:
                name_strategy.append("Названия стратегии в файле не обнаружено")
    print(name_strategy)
    return [result, name_strategy]


async def download_file(file_path: str, message: types.Message) -> str:
    """Скачать файл с серверов telegram.

    Скачивает в папку ./data/user_id/

    :param file_path: Путь к файлу на тг сервере.
    :param message: Сообщение от пользователя.
    :return: Путь к файлу на компьютере
    """
    if not os.path.exists(os.path.join(os.getcwd(), 'data')):
        os.mkdir(os.path.join(os.getcwd(), 'data'))
    if not os.path.exists(os.path.join(os.getcwd(), 'data', str(message.chat.id))):
        os.mkdir(os.path.join(os.getcwd(), 'data', str(message.chat.id)))

    extension_of_file = file_path.split('.')[-1]
    await bot.download_file(file_path, os.path.join(os.getcwd(), 'data', str(message.chat.id),
                                                    'report.' + extension_of_file))
    return os.path.join(os.getcwd(), 'data', str(message.chat.id), 'report.' + extension_of_file)


def get_func_to_parse(strategy_name: str):
    """Возвращает объект функции, которая будет обрабатывать комментарий определенного вида.

    :param strategy_name: Название стратегии в начале комментария.
    :return: function: Функция, которую можно вызвать в программе.
    """
    if strategy_name.find('dropsdetection') != -1:
        return parse_DropsDetection_comment
    elif strategy_name.find('joined sell') != -1:
        return parse_joined_sell
    elif strategy_name.find('moonstrike') != -1:
        return parse_moonstrike_comment
    elif strategy_name.find('moonshot') != -1:
        return parse_moonshot
    elif strategy_name.find('spreaddetection') != -1:
        return parse_StreadDetection
    elif strategy_name.find('emadetection') != -1:
        return parse_EMA_comment
    elif strategy_name.find('AutoDetection') != -1:
        return parse_moonstrike_comment
    elif strategy_name.find('pumpdetection') != -1:
        return parse_PuvpDetection
    elif strategy_name.find('hook long depth') != -1:
        return parse_hook_long_depth_comment
    elif strategy_name.find('trend') != -1:
        return parse_ichi_trend_bot
    elif strategy_name.find('waves') != -1:
        return parse_waves_comment
    else:
        return None


async def start_parse(file_path: str, message: types.Message, state: FSMContext, file_name: str) -> os.path:
    """ Главная функция обработки файла.

    Получает все заголовки и создаёт табличку из отчёта.

    :param file_path: Путь к файлу, который обрабатываем.
    :param message: Сообщение, полученное от пользователя.
    :param state: Состояние, в котором находится пользователь.
    :return: os.path: Путь к табличке, которая получена в результате обработки
    """
    await message.answer("Начинаю предварительный просмотр файла.", reply_markup=types.ReplyKeyboardRemove())
    file_on_pc = await download_file(file_path, message)
    if tarfile.is_tarfile(file_on_pc):
        tar_file = tarfile.TarFile(file_on_pc)
        chek_tar = [text_file for text_file in tar_file.getnames()]
        if len(chek_tar) > 1:
            await message.answer("В архиве может находиться только 1 файл формата TXT")
            logging.error(f"Ошибка в 172 строке файла main_parser.py. В архиве было передано {len(chek_tar)} файлов")
        with tarfile.TarFile(file_on_pc, 'r') as tar_file:
            file_on_pc = tar_file.extractfile(chek_tar[0])
    elif zipfile.is_zipfile(file_on_pc):
        zip_file = zipfile.ZipFile(file_on_pc)
        chek_zip = [text_file.filename for text_file in zip_file.infolist() ]
        if len(chek_zip) > 1:
            await message.answer("В архиве может находиться только 1 файл формата TXT")
            logging.error(f"Ошибка в 180 строке файла main_parser.py. В архиве было передано {len(chek_zip)} файлов")
        with zipfile.ZipFile(file_on_pc, 'r') as zip_file:
            file_on_pc = zip_file.extract(chek_zip[0], file_path)
    kostel = await get_headers(file_on_pc, message, state)
    strategies, name_strategies = kostel[0], kostel[1]
    await message.answer(f"Количество стратегий в файле: {len(set(name_strategies))}.\n"
                         f"Создаю таблицу")

    result = pd.DataFrame()

    with open(file_on_pc, 'r') as file:
        _ = file.readline()
        headers = [header.strip() for header in file.readline().strip().split('\t')]
        index_profite_abs = -1
        if "ProfitUSDT" in headers:
            index_profite_usdt = headers.index("ProfitUSDT")
        else:
            index_profite_usdt = headers.index("Profit USDT")
        if "ProfitAbs" in headers:
            index_profite_abs = headers.index("ProfitAbs")
        elif "Profit Abs" in headers:
            index_profite_abs = headers.index("Profit Abs")
        comment_index = headers.index('Comment')
        dct_pol, dct_otr, dct_profit_usdt, dct_profit_abs, res_dict = {}, {}, {}, {}, {"Название стратегии": [], "Монета": [], "Число положительных сделок": [],\
                                                              "Число отрицательных сделок": [], "Профит USDT": [],\
                                                              "Профит ABS (%)": []}
        for ind, line in enumerate(file.readlines()):
            line = line.strip().split('\t')
            try:
                strategy_parse = strategies[ind].lower()
            except IndexError:
                logging.warning(f"IndexError in start_parse function. {ind}/{len(strategies)}")
                continue
            try:
                if len(line) == 1:
                    continue
                if line[comment_index].split()[0] != 'Waves:':
                    #print(line)
                    if line[comment_index][:20].find("<") != -1:
                        key = line[0] + line[comment_index][line[comment_index].find("<"):line[comment_index].find(">") + 1]
                    elif line[comment_index][:50].find(" (") != -1:
                        key = line[0] + line[comment_index].split(" (")[0].split()[-1]
                    elif line[comment_index][:50].find("using strategy") != -1:
                        key = line[0] + line[comment_index].split("using strategy ")[0].split()[0]
                    else:
                        key = "Названия стратегии в файле не обнаружено"
                    if float(line[index_profite_usdt]) > 0:
                        dct_pol[key] = dct_pol.get(key, 0) + 1
                    else:
                        dct_otr[key] = dct_otr.get(key, 0) + 1
                    dct_profit_usdt[key] = dct_profit_usdt.get(key, 0) + float(line[index_profite_usdt])
                    if index_profite_abs != -1:
                        dct_profit_abs[key] = dct_profit_abs.get(key, 0) + float(line[index_profite_abs][:-2])
                else:
                    key = line[0] + line[7][line[7].find("<"):line[7].find(">") + 1]
                    if float(line[index_profite_usdt]) > 0:
                        dct_pol[key] = dct_pol.get(key, 0) + 1
                    else:
                        dct_otr[key] = dct_otr.get(key, 0) + 1
                    dct_profit_usdt[key] = dct_profit_usdt.get(key, 0) + float(line[index_profite_usdt])
                    if index_profite_abs != -1:
                        try:
                            dct_profit_abs[key] = dct_profit_abs.get(key, 0) + float(line[index_profite_abs])
                        except ValueError:
                            dct_profit_abs[key] = dct_profit_abs.get(key, 0) + float(line[index_profite_abs][:-2])


            except TypeError as error:
                logging.error(f"Error in 164 - 167 main_parser.py")
                logging.error(f"Error message: {error}")
            try:
                parsed_comment = get_func_to_parse(strategy_parse)(line[comment_index])
            except TypeError:
                print(strategy_parse)
                logging.error("Ошибка в 224 строке parsers/main_parser.py")
            try:
                tmp_df = pd.DataFrame([line + parsed_comment[0]], columns=headers + parsed_comment[1])
                tmp_df = tmp_df.fillna(value="-")
            except TypeError as error:
                logging.error(f"I don't know how to parse: '{strategy_parse}'")
                logging.error(f"Error message: {error}")
                continue
            except IndexError:
                continue
            except ValueError as error:
                logging.error(f"ValueError while parsing {strategy_parse}")
                logging.error(f"Error message: {error}")
                continue

            if result.empty:
                result = tmp_df
            else:
                try:
                    result = pd.concat([result, tmp_df], ignore_index=True, verify_integrity=True)
                except pd.errors.InvalidIndexError as error:
                    logging.error(f"Error while parsing {strategy_parse}.\n"
                                  f"Message: {error}")
                    continue
        for key, value in dct_pol.items():
            try:
                if key.find("<") != -1:
                    res_dict["Название стратегии"].append(key[key.find("<") + 1:key.find(">")])
                else:
                    res_dict["Название стратегии"].append(key.split()[1])
                res_dict["Монета"].append(key.split()[0])
                res_dict["Профит USDT"].append(dct_profit_usdt[key])
                if key in dct_profit_abs.keys():
                    res_dict["Профит ABS (%)"].append(dct_profit_abs[key])
                res_dict["Число положительных сделок"].append(value)
                res_dict["Число отрицательных сделок"].append(dct_otr.get(key, 0))
            except IndexError:
                 logging.error(f'{key} - {value}')
        profit_abs_mas = res_dict["Профит ABS (%)"]
        if len(profit_abs_mas) != len(res_dict["Название стратегии"]):
            del res_dict["Профит ABS (%)"]
            result_second = pd.DataFrame.from_dict(res_dict).sort_values(by=["Название стратегии", "Профит USDT"], ascending=False)
        else:
            result_second = pd.DataFrame.from_dict(res_dict).sort_values(by=["Название стратегии", "Профит USDT"], ascending=False)
        profit_list = result_second[["Профит USDT"]].values.tolist()
        indexs_per_value = {}
        info_abs_profit = []
        index_el = 2
        headers_result_second = result_second.columns
        max_pol_profit_abs, max_otr_profit_abs = "", ""
        max_pol_el, max_otr_el = "", ""
        for index in sorted(result_second["Название стратегии"].value_counts().to_dict().items(), key=lambda item: item[0], reverse=True):
            flag = -1
            try:
                max_pol_el = max(result_second["Число положительных сделок"][index_el - 2:index_el + index[1] - 3])
                max_otr_el = max(result_second["Число отрицательных сделок"][index_el - 2:index_el + index[1] - 3])
            except:
                print(result_second["Число положительных сделок"][index_el - 2:index_el + index[1] - 3])
                print(index)

            if "Профит ABS (%)" in headers_result_second:
                try:
                    max_pol_profit_abs = max(result_second["Профит ABS (%)"][index_el - 2:index_el + index[1] - 3])
                    max_otr_profit_abs = min(result_second["Профит ABS (%)"][index_el - 2:index_el + index[1] - 3])
                except:
                    print(result_second["Профит ABS (%)"][index_el - 2:index_el + index[1] - 3])

            for i in range(index_el, index_el + index[1] - 2):
                if profit_list[i][0] < 0:
                    indexs_per_value[index[0]] = [index_el, i + 2, index[1] + index_el - 1,
                                                  profit_list[index_el - 2][0], profit_list[i][0],
                                                  profit_list[index_el + index[1] - 3][0]]
                    flag = 1
                    index_el += index[1]
                    break
            if flag == -1:
                indexs_per_value[index[0]] = [index_el, index[1] + index_el - 1, profit_list[index_el - 2][0],
                                              profit_list[index[1] + index_el - 3][0]]
                index_el += index[1]

            if type(max_pol_el) != str:
                indexs_per_value[index[0]].extend([max_pol_el, max_otr_el])

            if type(max_pol_profit_abs) != str:
                info_abs_profit.append([max_pol_profit_abs, max_otr_profit_abs])

        list_math = get_mean_plot(result)
        result_third, info_math_list = list_math[0], list_math[1]
    await create_xlsx(result, result_second, result_third, message, file_name,
                      indexs_per_value, info_abs_profit, info_math_list)


def create_diagram_xlsx(xlsxname, size, index_per_value: dict, info_abs_profit: list, info_math_list: dict) -> None:
    workbook = openpyxl.load_workbook(xlsxname)
    worksheert_profit = workbook['Профит']
    worksheert_diagrams = workbook.create_sheet(title="Диаграммы")
    worksheert_math = workbook['Мат.Ожидания']

    values1 = Reference(worksheert_profit, min_col=5, min_row=1, max_col=5, max_row=size + 1)
    monet_name_1 = Reference(worksheert_profit, min_col=2, min_row=2, max_col=2, max_row=size + 1)
    values2 = Reference(worksheert_profit, min_col=3, min_row=1, max_col=4, max_row=size + 1)

    ''' Диаграммы на листе профит '''
    chart1 = BarChart()
    chart1.type = "bar"
    chart1.width = 15
    chart1.height = 0.5 * (size + 1)
    #chart1.y_axis.title = 'Величина профита'
    chart1.add_data(values1, titles_from_data=True)
    chart1.set_categories(monet_name_1)

    chart2 = BarChart()
    chart2.type = "bar"
    chart2.width = 15
    chart2.height = 0.5 * (size + 1)
    chart2.y_axis.title = 'Количество сделок'
    chart2.title = 'Статистика сделок'
    chart2.add_data(values2, titles_from_data=True)
    chart2.set_categories(monet_name_1)

    worksheert_diagrams.add_chart(chart1, "A1")
    worksheert_diagrams.add_chart(chart2, "J1")

    flag = 19
    for key, value in info_math_list.items():
        if len(value) == 4:

            categories_math = Reference(worksheert_math, min_col=3, min_row=value[0], max_col=3, max_row=value[1])
            values_math = Reference(worksheert_math, min_col=4, min_row=value[0], max_col=4, max_row=value[1])
            chart3 = BarChart()
            chart3.type = "bar"
            chart3.width = 15
            chart3.height = 0.5 * (value[1] - value[0] + 1)
            chart3.y_axis.title = 'Мат. ожидание'
            chart3.title = f'Мат. ожидание {key}'
            chart3.add_data(values_math)
            chart3.set_categories(categories_math)
            worksheert_diagrams.add_chart(chart3, f"{get_column_letter(flag)}1")

            data_bar_rule_otr = DataBarRule(start_type="num",
                                            start_value=value[2],
                                            end_type="num",
                                            end_value=value[3],
                                            color="B22222")
            worksheert_math.conditional_formatting.add(f"D{value[0]}:D{value[1]}", data_bar_rule_otr)

        elif len(value) == 5:

            categories_math = Reference(worksheert_math, min_col=3, min_row=value[0], max_col=3, max_row=value[2])
            values_math = Reference(worksheert_math, min_col=4, min_row=value[0], max_col=4, max_row=value[2])
            chart3 = BarChart()
            chart3.type = "bar"
            chart3.width = 15
            chart3.height = 0.5 * (value[2] - value[0] + 1)
            chart3.y_axis.title = 'Мат. ожидание'
            chart3.title = f'Мат. ожидание {key}'
            chart3.add_data(values_math)
            chart3.set_categories(categories_math)
            worksheert_diagrams.add_chart(chart3, f"{get_column_letter(flag)}1")

            data_bar_rule_otr = DataBarRule(start_type="num",
                                            start_value=0,
                                            end_type="num",
                                            end_value=value[3],
                                            color="B22222")
            worksheert_math.conditional_formatting.add(f"D{value[0]}:D{value[1]}", data_bar_rule_otr)

            data_bar_rule_pol = DataBarRule(start_type="num",
                                            start_value=0,
                                            end_type="num",
                                            end_value=value[4],
                                            color="002aff")
            worksheert_math.conditional_formatting.add(f"D{value[1]}:D{value[2]}", data_bar_rule_pol)

        flag += 10

    for index, value in enumerate(index_per_value.values()):
        try:
            if len(value) == 6:
                data_bar_rule_pol = DataBarRule(start_type="num",
                                                start_value=0,
                                                end_type="num",
                                                end_value=value[4],
                                                color="002aff")
                worksheert_profit.conditional_formatting.add(f"C{value[0]}:C{value[1]}", data_bar_rule_pol)

                data_bar_rule_otr = DataBarRule(start_type="num",
                                                start_value=0,
                                                end_type="num",
                                                end_value=value[5],
                                                color="B22222")
                worksheert_profit.conditional_formatting.add(f"D{value[0]}:D{value[1]}", data_bar_rule_otr)

                data_bar_rule_pol_profit_usdt = DataBarRule(start_type="num",
                                                            start_value=value[3],
                                                            end_type="num",
                                                            end_value=value[2],
                                                            color="002aff")
                worksheert_profit.conditional_formatting.add(f"E{value[0]}:E{value[1]}", data_bar_rule_pol_profit_usdt)


            else:
                data_bar_rule_pol = DataBarRule(start_type="num",
                                                start_value=0,
                                                end_type="num",
                                                end_value=value[6],
                                                color="002aff")
                worksheert_profit.conditional_formatting.add(f"C{value[0]}:C{value[2]}", data_bar_rule_pol)

                data_bar_rule_otr = DataBarRule(start_type="num",
                                                start_value=0,
                                                end_type="num",
                                                end_value=value[7],
                                                color="B22222")
                worksheert_profit.conditional_formatting.add(f"D{value[0]}:D{value[2]}", data_bar_rule_otr)

                data_bar_rule_pol_profit_usdt = DataBarRule(start_type="num",
                                                            start_value=0.0,
                                                            end_type="num",
                                                            end_value=value[3],
                                                            color='002aff')
                worksheert_profit.conditional_formatting.add(f"E{value[0]}:E{value[1] - 1}", data_bar_rule_pol_profit_usdt)

                data_bar_rule_otr_profit_usdt = DataBarRule(start_type="num",
                                                            start_value=value[4],
                                                            end_type="num",
                                                            end_value=value[5],
                                                            color="B22222")
                worksheert_profit.conditional_formatting.add(f"E{value[1]}:E{value[2]}",data_bar_rule_otr_profit_usdt)

            if len(info_abs_profit) >= 1:
                if len(value) == 8:
                    data_bar_rule_pol_profit_abs = DataBarRule(start_type="num",
                                                            start_value=0,
                                                            end_type="num",
                                                            end_value=info_abs_profit[index][0],
                                                            color="002aff")
                    worksheert_profit.conditional_formatting.add(f"F{value[0]}:F{value[1] - 1}", data_bar_rule_pol_profit_abs)

                    data_bar_rule_otr_profit_abs = DataBarRule(start_type="num",
                                                           start_value=0,
                                                           end_type="num",
                                                           end_value=info_abs_profit[index][1],
                                                           color="B22222")
                    worksheert_profit.conditional_formatting.add(f"F{value[1]}:F{value[2]}", data_bar_rule_otr_profit_abs)
                else:
                    try:
                        data_bar_rule_pol_profit_abs = DataBarRule(start_type="num",
                                                               start_value=0,
                                                               end_type="num",
                                                               end_value=info_abs_profit[index][0],
                                                               color="002aff")
                        worksheert_profit.conditional_formatting.add(f"F{value[0]}:F{value[1] - 1}",
                                                                    data_bar_rule_pol_profit_abs)
                    except TypeError:
                        print(f"F{value[0]}:F{value[1] - 1}")
        except:
            logging.error("Заебало!")



    workbook.save(xlsxname)


async def create_xlsx(data1: pd.DataFrame, data2: pd.DataFrame, data3: pd.DataFrame, message: types.Message,
                      file_name: str, info_math: dict, info_abs_profit: list, info_math_list: list) -> None:
    """Создаёт табличку из обработанного отчёта и отправляет пользователю.

    :param data1: Обработанные данные.
    :param message: Сообщение от пользователя.
    """
    with ExcelWriter(os.path.join(PATH_TO_RESULT_REPORT, f'{message.chat.id}.xlsx'), mode="w") as writer:
        data1.to_excel(writer, sheet_name="Отчет", index=False)
        data2.to_excel(writer, sheet_name="Профит", index=False)
        data3.to_excel(writer, sheet_name="Мат.Ожидания", index=False)
    create_diagram_xlsx(os.path.join(PATH_TO_RESULT_REPORT, f'{message.chat.id}.xlsx'), len(data2.index),
                        info_math, info_abs_profit, info_math_list)
    await message.answer(f"Отправляю обработанный отчёт...", reply_markup=types.ReplyKeyboardRemove())
    flag = False
    start_time = time.time()

    # Ждем 20 сек на создание файла
    while not flag and time.time() - start_time < 20:
        try:
            file = input_file.InputFile(os.path.join(PATH_TO_RESULT_REPORT, f'{message.chat.id}.xlsx'),
                                        filename=f'{".".join(file_name.split(".")[:-1])}_parsed.xlsx')

            await message.answer_document(file)
            flag = True
        except Exception:
            continue
    await message.answer(f"Нажмите кнопку /parse, если хотите распарсить еще один отчет."
                         f" Также вы можете загрузить исторические данные, введя ккоманду /load. ",
                         reply_markup=types.ReplyKeyboardRemove())

